#!/usr/bin/env python3
"""Convert MCQsim xlsx deliverables to Avi-style benchmark CSV bundles."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT_DIR = REPO_ROOT / "benchmark/Simulation_results/MCQsim_results"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "benchmark/Simulation_results/MCQsim_results"
DEFAULT_MIN_MW = 7.0

EVENT_COL_MAP = {
    "EventID":       "event_id",
    "HypoTimeSec":   "t0_s",
    "HypoTimeYears": "t0_year",
    "Moment":        "m0_Nm",
    "Magnitude":     "mw",
    "RuptArea":      "area_m2",
    "RuptDuration":  "dt_s",
    "MaxSlip":       "max_slip_m",
    "HypoLoc_E":     "x_NZTM_m",
    "HypoLoc_N":     "y_NZTM_m",
    "HypoLoc_Z":     "z_NZTM_m",
    "SouthEnd_N":    "sbound_NZTM_m",
    "NorthEnd_N":    "nbound_NZTM_m",
}

SITE_COL_MAP = {
    "ID":          "event_id",
    "time (sec)":  "t0_s",
    "M_w":         "Mw",
    "slip":        "Slip_m",
    "rake":        "Rake_deg",
}

EVENT_FIELDS = list(EVENT_COL_MAP.values())
SITE_FIELDS  = list(SITE_COL_MAP.values())

SCENARIOS = [
    {
        "prefix":     "alpine_planar",
        "event_xlsx": "AlpineFault_Planar_45kyr_catalog_oz.xlsx",
        "site_xlsx":  "AlpineFaut_Planar_sites_1to3_oz.xlsx",
    },
    {
        "prefix":     "alpine_varying_dip",
        "event_xlsx": "AlpineFault_VariDip_45kyr_catalog_oz.xlsx",
        "site_xlsx":  "AlpineFaut_VariDip_sites_1to3_oz.xlsx",
    },
]

SITE_SHEET_ORDER = [1, 2, 3]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert MCQsim xlsx files to benchmark CSV bundles.")
    parser.add_argument("--input-dir",  type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--min-mw",     type=float, default=DEFAULT_MIN_MW)
    return parser.parse_args()


def xlsx_to_rows(path: Path, col_map: dict[str, str], required_fields: list[str]) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    col_indices = {col_map[h]: raw_headers.index(h) for h in col_map if h in raw_headers}
    missing = [f for f in required_fields if f not in col_indices]
    if missing:
        raise ValueError(f"{path.name}: missing mapped columns {missing}")
    rows = []
    for raw in rows_iter:
        if all(raw[col_indices[f]] is None for f in required_fields):
            continue
        rows.append({f: raw[col_indices[f]] for f in required_fields})
    wb.close()
    return rows


def xlsx_site_rows(path: Path) -> dict[int, list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[int, list[dict]] = {}
    for site_n, sheet_name in zip(SITE_SHEET_ORDER, wb.sheetnames):
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        col_indices = {SITE_COL_MAP[h]: raw_headers.index(h) for h in SITE_COL_MAP if h in raw_headers}
        missing = [f for f in SITE_FIELDS if f not in col_indices]
        if missing:
            raise ValueError(f"{path.name} sheet '{sheet_name}': missing columns {missing}")
        site_rows = []
        for raw in rows_iter:
            if raw[col_indices["event_id"]] is None:
                continue
            site_rows.append({f: raw[col_indices[f]] for f in SITE_FIELDS})
        result[site_n] = site_rows
    wb.close()
    return result


def write_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    input_dir  = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()

    for scenario in SCENARIOS:
        prefix     = scenario["prefix"]
        event_path = input_dir / scenario["event_xlsx"]
        site_path  = input_dir / scenario["site_xlsx"]

        for p in (event_path, site_path):
            if not p.exists():
                raise FileNotFoundError(p)

        event_rows = xlsx_to_rows(event_path, EVENT_COL_MAP, EVENT_FIELDS)
        # SouthEnd_N / NorthEnd_N are delivered in km; benchmark format expects metres
        for r in event_rows:
            r["sbound_NZTM_m"] = float(r["sbound_NZTM_m"]) * 1e3
            r["nbound_NZTM_m"] = float(r["nbound_NZTM_m"]) * 1e3
        event_rows = [r for r in event_rows if float(r["mw"]) > args.min_mw]
        if not event_rows:
            raise RuntimeError(f"No events above Mw {args.min_mw} for {prefix}")

        site_data = xlsx_site_rows(site_path)

        write_csv(output_dir / f"{prefix}_event_info.csv", event_rows, EVENT_FIELDS)
        print(output_dir / f"{prefix}_event_info.csv")
        for site_n, rows in site_data.items():
            out = output_dir / f"{prefix}_site{site_n}_event_info.csv"
            write_csv(out, rows, SITE_FIELDS)
            print(out)


if __name__ == "__main__":
    main()
